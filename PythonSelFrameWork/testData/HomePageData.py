import openpyxl

class HomePageData:

    #test_homepage_data = [{"firstName" : "Vetri", "email" : "vkp@gmail.com", "gender" : "Male"},
    #                     {"firstName" : "Manjula", "email" : "mvkp@hotmail.com", "gender" : "Female"}]

    @staticmethod
    def gettestData():
        book = openpyxl.load_workbook("C:\\Users\\vetri\\PythonDemo.xlsx")
        sheet = book.active

        rows = sheet.max_row
        columns = sheet.max_column
        listdata = []

        for r in range(2, rows + 1):
            dictonary = {}
            for c in range(2, columns + 1):
                dictonary[sheet.cell(1, c).value] = sheet.cell(r, c).value
            listdata.append(dictonary)
        return listdata